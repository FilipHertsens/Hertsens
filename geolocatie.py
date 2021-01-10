from shapely.geometry import Point, Polygon
from geopy.geocoders import Nominatim
from openpyxl import load_workbook
from openpyxl.styles.colors import Color
from pprint import pprint

def GetPlaces(file='data/Tuinonderhoud lijst adressen.xlsx'):
    polys = []
    pol = {}
    rhead = 1
    cstart = 1
    r = 1
    c = 1
    wb = load_workbook(filename=file)
    blad = wb['Blad1']
    while blad.cell(row = rhead, column = c).value != 'Number':
        rhead += 1
    while blad.cell(row = rhead+r,column=cstart).value != None:
        while blad.cell(row = rhead, column = cstart+c-1).value != None:
            colname = blad.cell(row = rhead, column = c).value
            if colname == 'Color':
                cel = blad.cell(row = rhead+r, column = c).fill.start_color.index
            else:
                cel = blad.cell(row = rhead+r, column = c).value
            pol[colname] = cel
            c += 1
        c = 1
        polys.append(pol)
        pol = {}
        r += 1
    #pprint(polys)
    return polys

def ExDatSms(dat):
    link = dat.split(' ',1)[0]
    cor = link.split('?q=',1)[1].split('%2c',1)
    cor = cor[0]+','+cor[1]
    poly = InPoly(cor)
    adres = CooToAdress(cor)
    time = dat.split('V:A,',1)[1].split(' S:')[0]
    unit = dat.split(',')[3]

    return link, cor, poly, adres, time, unit

def InPoly(co):
    if isinstance(co,dict):
        co = str(co['latitude'])+', '+str(co['longitude'])

    places = GetPlaces()
    for plase in places:

        if plase['coord 1'] != None and plase['coord 2'] != None and plase['coord 3'] != None and plase['coord 4'] != None:
            a = eval(plase['coord 1'])
            b = eval(plase['coord 2'])
            c = eval(plase['coord 3'])
            d = eval(plase['coord 4'])
            p = a,b,c,d
            poly = Polygon(p)
            #print plase['Street']
            if Point(eval(co)).within(poly):
                if plase['Nickname'] != None:
                    return plase['Nickname']
                else:
                    adres = plase['Street']+''+plase['City']
                    return adres


    return None

def CooToAdress(co):
    geolocator = Nominatim(user_agent="testing")
    location = geolocator.reverse(co)
    return location.address


'''
sms = 'http://maps.google.com/maps?q=51.193831%2c004.317427   V:A,2020-12-02 07:31:59 S:008km/h,Bat:6ACC:off ,7201075959,S20G05;p:20620, l:1251, c:4746'
t = ExDatSms(sms)
for i in t:
    print(i)  
'''

